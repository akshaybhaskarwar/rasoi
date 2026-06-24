#!/usr/bin/env python3
"""
Bulk import recipes for a single user from an Excel sheet.

Usage
-----
    # Generate a starter template the operator can fill in.
    python scripts/import_recipes.py --generate-template recipes_template.xlsx

    # Dry-run a filled-in file — validate everything, write nothing.
    python scripts/import_recipes.py recipes.xlsx \
        --user-email nalini@example.com --dry-run

    # Insert for real. Recipes are private (is_published=False) by
    # default; pass --publish to push them straight to the community
    # feed instead.
    python scripts/import_recipes.py recipes.xlsx \
        --user-email nalini@example.com
    python scripts/import_recipes.py recipes.xlsx \
        --user-email nalini@example.com --publish

Excel schema (one row per recipe, sheet name doesn't matter)
------------------------------------------------------------

Columns (header row, case-insensitive, extra columns ignored):

    title                ✱ required, free text
    chef_name              optional, defaults to the user's name
    story                  optional, family/heritage note
    ingredients          ✱ required, ONE INGREDIENT PER LINE inside the
                           cell, format `name | quantity | unit`
                           e.g.  Rava | 1 | cup
                                 Curd | 3/4 | cup
    instructions         ✱ required, ONE STEP PER LINE inside the cell
    tags                   optional, comma-separated, e.g. `breakfast, vegetarian`
    servings               optional integer, defaults to 4
    prep_time_minutes      optional integer
    cook_time_minutes      optional integer

Photos are deliberately not part of the import. The recipe owner adds
them later from the website via the standard edit-recipe flow.

Why a script (and not an admin endpoint)
----------------------------------------

This is a one-off operator task — Akshay sitting at a terminal,
seeding 10-30 recipes for a family member who can then claim and
photo-tag them. A CLI script is faster to write, easier to dry-run,
and doesn't introduce an admin-auth surface area.

The script writes the SAME document shape that the live
`POST /api/recipes` endpoint writes (see backend/recipes.py around the
recipe_doc dict). The only differences are:
  - No inventory-match step (cheap but adds N Mongo queries per row;
    skipping makes a 30-recipe import O(1) instead of O(N·M))
  - No translation step (would require the Google Translate API; the
    frontend already falls back to ingredient_name when locale fields
    are missing, so the recipe still works in every language)

Both can be filled in later by editing the recipe via the website if
the household values inventory-stock-badges and Marathi/Hindi labels
on the imported recipes — and most operators won't, so saving the API
calls is the right default.
"""
from __future__ import annotations

import argparse
import asyncio
import os
import re
import sys
import uuid
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable

# Resolve paths so the script can be invoked from anywhere.
SCRIPT_DIR = Path(__file__).resolve().parent
BACKEND_DIR = SCRIPT_DIR.parent
# Surface backend imports (models live under backend/) for any future
# additions; right now we only need env loading from the backend dir.
sys.path.insert(0, str(BACKEND_DIR))

try:
    from dotenv import load_dotenv
    load_dotenv(BACKEND_DIR / ".env")
except ImportError:
    # python-dotenv is in requirements.txt for the running server, so
    # it should always be available — but don't crash a help-only run.
    pass


# Fractional-quantity strings users routinely type into Excel ("3/4
# cup", "1 1/2 tsp"). Pure float() can't parse them; this helper does.
_FRACTION_RE = re.compile(r"^\s*(\d+)?\s*(\d+)\s*/\s*(\d+)\s*$")


def parse_quantity(raw: str) -> float:
    """Convert a quantity cell to a float.

    Accepts integers (`2`), decimals (`1.5`), simple fractions (`3/4`),
    and mixed numbers (`1 1/2`). Anything else raises ValueError so
    the caller can flag a bad row.
    """
    s = str(raw).strip()
    if not s:
        raise ValueError("quantity is empty")

    # Plain numeric path.
    try:
        return float(s)
    except ValueError:
        pass

    m = _FRACTION_RE.match(s)
    if m:
        whole, num, denom = m.groups()
        denom_i = int(denom)
        if denom_i == 0:
            raise ValueError(f"zero denominator in {raw!r}")
        return (int(whole) if whole else 0) + int(num) / denom_i

    raise ValueError(f"can't parse quantity {raw!r}")


def parse_ingredient_line(line: str) -> dict:
    """Parse one `name | qty | unit` line into a RecipeIngredient dict."""
    parts = [p.strip() for p in line.split("|")]
    if len(parts) != 3:
        raise ValueError(
            f"ingredient line must be `name | quantity | unit` (got {line!r})"
        )
    name, qty_raw, unit = parts
    if not name:
        raise ValueError("ingredient name is blank")
    if not unit:
        raise ValueError(f"unit is blank in {line!r}")
    return {
        "ingredient_name": name,
        "inventory_item_id": None,
        "quantity": parse_quantity(qty_raw),
        "unit": unit.lower(),
        "name_en": name,   # bare fallback so locale-aware UI still has text
        "name_mr": None,   # filled in later if/when user edits via website
        "name_hi": None,
    }


def split_lines(cell: Any) -> list[str]:
    """Split a multi-line cell into trimmed non-empty lines."""
    if cell is None:
        return []
    return [ln.strip() for ln in str(cell).splitlines() if ln.strip()]


def split_tags(cell: Any) -> list[str]:
    if cell is None:
        return []
    return [t.strip() for t in str(cell).split(",") if t.strip()]


def to_int_or_none(cell: Any) -> int | None:
    if cell is None or str(cell).strip() == "":
        return None
    try:
        return int(float(str(cell).strip()))
    except ValueError:
        return None


def normalize_header(name: Any) -> str:
    return str(name or "").strip().lower().replace(" ", "_")


# Columns the operator can put in any order. Unknown columns are
# silently ignored so people can keep notes / scratch columns in the
# workbook without breaking the parser.
KNOWN_COLUMNS = {
    "title", "chef_name", "story", "ingredients", "instructions",
    "tags", "servings", "prep_time_minutes", "cook_time_minutes",
}


def parse_row(headers: list[str], row: tuple, row_num: int) -> dict:
    """Build a recipe dict from one Excel row.

    Raises ValueError with a human-readable message on bad input.
    """
    data: dict[str, Any] = {}
    for col_idx, header in enumerate(headers):
        if header in KNOWN_COLUMNS and col_idx < len(row):
            data[header] = row[col_idx]

    title = str(data.get("title") or "").strip()
    if not title:
        raise ValueError("title is blank")

    ing_lines = split_lines(data.get("ingredients"))
    if not ing_lines:
        raise ValueError("ingredients column is empty")

    ingredients: list[dict] = []
    for i, line in enumerate(ing_lines, 1):
        try:
            ingredients.append(parse_ingredient_line(line))
        except ValueError as e:
            raise ValueError(f"ingredient #{i}: {e}") from None

    step_lines = split_lines(data.get("instructions"))
    if not step_lines:
        raise ValueError("instructions column is empty")

    instructions = [
        {"step_number": i, "instruction": s, "duration_minutes": None}
        for i, s in enumerate(step_lines, 1)
    ]

    return {
        "title": title,
        "chef_name": (str(data.get("chef_name") or "").strip() or None),
        "story": (str(data.get("story") or "").strip() or None),
        "ingredients": ingredients,
        "instructions": instructions,
        "tags": split_tags(data.get("tags")),
        "servings": to_int_or_none(data.get("servings")) or 4,
        "prep_time_minutes": to_int_or_none(data.get("prep_time_minutes")),
        "cook_time_minutes": to_int_or_none(data.get("cook_time_minutes")),
    }


def build_recipe_doc(parsed: dict, user: dict, household_id: str, publish: bool) -> dict:
    """Mirror the document shape that POST /api/recipes writes.

    Important: keep this in lockstep with recipes.py:create_recipe so
    list/detail endpoints render imported recipes identically to
    ones created via the website.
    """
    now_iso = datetime.now(timezone.utc).isoformat()
    return {
        "id": str(uuid.uuid4()),
        "household_id": household_id,
        "created_by": user.get("id"),
        "created_by_name": user.get("name"),
        "created_by_email": user.get("email"),
        "title": parsed["title"],
        "chef_name": parsed["chef_name"] or user.get("name"),
        "story": parsed["story"],
        "ingredients": parsed["ingredients"],
        "instructions": parsed["instructions"],
        "tags": parsed["tags"],
        "servings": parsed["servings"],
        "prep_time_minutes": parsed["prep_time_minutes"],
        "cook_time_minutes": parsed["cook_time_minutes"],
        "photo_url": None,
        "is_published": publish,
        "likes": 0,
        "created_at": now_iso,
        "updated_at": now_iso,
    }


# ---------------------------------------------------------------------------
# Template generator — keeps the schema docs and the file in lockstep.
# ---------------------------------------------------------------------------

def generate_template(out_path: Path) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        sys.exit(
            "openpyxl not installed. Run: pip install openpyxl\n"
            "(or `pip install -r backend/requirements.txt` after pulling)"
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "recipes"

    headers = [
        "title", "chef_name", "story", "ingredients", "instructions",
        "tags", "servings", "prep_time_minutes", "cook_time_minutes",
    ]
    ws.append(headers)

    # Header styling — bold + light fill so the operator sees clearly
    # where the schema ends and their data begins.
    fill = PatternFill("solid", fgColor="FFEED8")
    bold = Font(bold=True)
    for cell in ws[1]:
        cell.font = bold
        cell.fill = fill
        cell.alignment = Alignment(vertical="center")

    example = [
        "Rava Dhokla",
        "Nalini Wargantiwar",
        "A quick steamed snack from western Maharashtra — a family favourite for Sunday breakfast.",
        "Rava | 1 | cup\nCurd | 3/4 | cup\nGreen chili | 1 | piece\nGinger | 1 | tsp\nEno fruit salt | 1 | tsp\nSalt | 1 | tsp\nOil | 1 | tbsp",
        "Mix rava and curd, rest the batter for 15 minutes.\nAdd chili paste, ginger, salt; whisk smooth.\nFold in eno just before steaming.\nPour into a greased thali and steam 12 minutes.\nTemper with mustard seeds + curry leaves, garnish coriander.",
        "breakfast, vegetarian",
        4,
        15,
        15,
    ]
    ws.append(example)

    # Column widths tuned so a screen-shot or print stays readable.
    widths = {"A": 24, "B": 22, "C": 40, "D": 38, "E": 50, "F": 22, "G": 10, "H": 12, "I": 12}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # Multi-line cells need wrap + top-aligned content to render the
    # newlines instead of running everything onto one row.
    for row in ws.iter_rows(min_row=2, max_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


# ---------------------------------------------------------------------------
# Mongo import — talks to the same DB the API server uses.
# ---------------------------------------------------------------------------

async def run_import(xlsx_path: Path, user_email: str, *, dry_run: bool, publish: bool) -> int:
    try:
        from openpyxl import load_workbook
    except ImportError:
        sys.exit(
            "openpyxl not installed. Run: pip install openpyxl\n"
            "(or `pip install -r backend/requirements.txt` after pulling)"
        )

    mongo_url = os.environ.get("MONGO_URL")
    db_name = os.environ.get("DB_NAME")
    if not mongo_url or not db_name:
        sys.exit(
            "MONGO_URL and DB_NAME must be set (load backend/.env or "
            "export them in the shell)."
        )

    from motor.motor_asyncio import AsyncIOMotorClient
    client = AsyncIOMotorClient(mongo_url)
    db = client[db_name]

    user = await db.users.find_one(
        {"email": user_email}, {"_id": 0, "hashed_password": 0}
    )
    if not user:
        sys.exit(f"no user with email {user_email!r}")
    household_id = user.get("active_household")
    if not household_id:
        sys.exit(
            f"user {user_email!r} has no active household — they need to "
            "create or join a kitchen on the website first."
        )

    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    rows: Iterable[tuple] = ws.iter_rows(values_only=True)

    try:
        header_row = next(rows)
    except StopIteration:
        sys.exit(f"{xlsx_path} has no header row")
    headers = [normalize_header(h) for h in header_row]

    parsed_docs: list[dict] = []
    errors: list[str] = []
    skipped: list[str] = []

    # row_num starts at 2 because the header is row 1 and Excel's
    # row references are 1-based — keeping the math straight makes
    # error messages copy-pasteable into the spreadsheet.
    for row_num, row in enumerate(rows, start=2):
        # Skip rows that are entirely empty so trailing blank lines in
        # the workbook don't get reported as errors.
        if all(cell is None or str(cell).strip() == "" for cell in row):
            continue
        try:
            parsed = parse_row(headers, row, row_num)
        except ValueError as e:
            errors.append(f"row {row_num}: {e}")
            continue
        parsed_docs.append((row_num, parsed))

    print(f"\nparsed {len(parsed_docs)} recipe(s) from {xlsx_path.name}")
    if errors:
        print(f"\n{len(errors)} row(s) had errors:")
        for e in errors:
            print(f"  ✗ {e}")
    else:
        print("  (no validation errors)")

    if dry_run:
        print("\n--dry-run set: not inserting. Re-run without --dry-run to commit.")
        if errors:
            return 1
        return 0

    if errors:
        sys.exit(
            "\nrefusing to insert while there are validation errors. "
            "Fix the rows above and re-run."
        )

    inserted_titles: list[str] = []
    for row_num, parsed in parsed_docs:
        # De-dup by (household_id, lowercase title) so re-running the
        # script after a partial failure doesn't double-add. A duplicate
        # is treated as "already there, skip" — not an error.
        existing = await db.user_recipes.find_one({
            "household_id": household_id,
            "title": {"$regex": f"^{re.escape(parsed['title'])}$", "$options": "i"},
        })
        if existing:
            skipped.append(f"row {row_num}: {parsed['title']!r} already exists")
            continue

        doc = build_recipe_doc(parsed, user, household_id, publish)
        await db.user_recipes.insert_one(doc)
        inserted_titles.append(parsed["title"])

    print(f"\ninserted {len(inserted_titles)} recipe(s) for {user_email}:")
    for title in inserted_titles:
        print(f"  ✓ {title}")
    if skipped:
        print(f"\nskipped {len(skipped)} (already in DB):")
        for s in skipped:
            print(f"  - {s}")

    print(
        f"\ndone. is_published={publish} — "
        f"{'recipes are live in the community feed.' if publish else 'recipes are private until owner publishes.'}"
    )
    return 0


# ---------------------------------------------------------------------------

def main() -> int:
    p = argparse.ArgumentParser(
        description="Bulk import recipes for one user from Excel.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    p.add_argument(
        "xlsx",
        nargs="?",
        type=Path,
        help="Path to the recipes .xlsx (omit when --generate-template).",
    )
    p.add_argument(
        "--user-email",
        help="Email of the user the recipes will be attributed to.",
    )
    p.add_argument(
        "--dry-run",
        action="store_true",
        help="Parse + validate but write nothing to MongoDB.",
    )
    p.add_argument(
        "--publish",
        action="store_true",
        help="Set is_published=True so recipes land on the community feed.",
    )
    p.add_argument(
        "--generate-template",
        type=Path,
        metavar="PATH",
        help="Write a starter Excel template (with one example row) to PATH and exit.",
    )
    args = p.parse_args()

    if args.generate_template:
        generate_template(args.generate_template)
        print(f"wrote template → {args.generate_template}")
        return 0

    if not args.xlsx:
        p.error("xlsx path is required (or pass --generate-template)")
    if not args.xlsx.exists():
        p.error(f"{args.xlsx} does not exist")
    if not args.user_email:
        p.error("--user-email is required for import")

    return asyncio.run(run_import(
        args.xlsx,
        args.user_email,
        dry_run=args.dry_run,
        publish=args.publish,
    ))


if __name__ == "__main__":
    raise SystemExit(main())
